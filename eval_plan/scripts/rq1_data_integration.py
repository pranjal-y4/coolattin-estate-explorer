#!/usr/bin/env python3
"""
eval_plan/scripts/rq1_data_integration.py

RQ1 evidence, per EVAL_RQ1_data_integration.md. Read-only / dry-run only —
does NOT write to coolattin.db and does NOT run a real full_ingest, because a
prior dry run showed the live VRTI census SPARQL endpoint currently returns 0
rows for Wicklow, which means a real rebuild right now would overwrite the
existing 8,033-row census_record table with ~1,462 rows. See
RQ1_data_integration_results.md for the full explanation before ever running
`full_ingest.run_full_ingest(dry_run=False)` against this database.

Produces:
  1. Ingestion completeness: raw source row counts vs currently-loaded DB counts.
  2. A dry-run of full_ingest (safe — writes nothing) to see what a rebuild
     would produce right now, for comparison against the current DB state.
  3. Townland alignment coverage, scoped to the ~152 estate townlands.

Run from repo root: venv/bin/python3 eval_plan/scripts/rq1_data_integration.py
"""
from __future__ import annotations

import json
import logging
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

# Silence the per-townland INFO logs from vrti_sparql during the dry run —
# we only want the summary stats, not 152 lines of KG fetch logs.
logging.getLogger("backend.integrations.vrti_sparql").setLevel(logging.WARNING)
logging.getLogger("backend.jobs.full_ingest").setLevel(logging.WARNING)

REPO_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")


def raw_source_counts() -> dict:
    import openpyxl

    counts = {}

    def _csv_rows(path):
        with open(path, encoding="utf-8") as f:
            return sum(1 for _ in f) - 1  # minus header

    counts["unified_processed.csv"] = _csv_rows(
        os.path.join(REPO_ROOT, "frontend/static/data/unified_processed.csv")
    )
    counts["unified_census.csv"] = _csv_rows(
        os.path.join(REPO_ROOT, "frontend/static/data/unified_census.csv")
    )
    counts["wicklow-census-data.csv"] = _csv_rows(
        os.path.join(REPO_ROOT, "frontend/static/data/wicklow-census-data.csv")
    )

    wb = openpyxl.load_workbook(
        os.path.join(REPO_ROOT, "frontend/static/data/workhouse_data_final.xlsx")
    )
    wh_total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        rows = sum(1 for row in ws.iter_rows(min_row=2) if any(c.value is not None for c in row))
        counts[f"workhouse_data_final.xlsx[{name}]"] = rows
        wh_total += rows
    counts["workhouse_data_final.xlsx[TOTAL]"] = wh_total

    with open(os.path.join(REPO_ROOT, "frontend/static/data/townlands.json"), encoding="utf-8") as f:
        tj = json.load(f)
    counts["townlands.json[features]"] = len(tj.get("features", []))

    return counts


def loaded_db_counts(conn) -> dict:
    tables = ["unified_record", "source_mentions", "census_record", "clearances_record", "townland"]
    return {t: conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0] for t in tables}


def alignment_coverage(conn) -> dict:
    return dict(
        conn.execute(
            """
            SELECT
              COUNT(DISTINCT t.id) AS total_matched_townland_rows,
              SUM(CASE WHEN t.osm_id IS NOT NULL AND t.osm_id<>'' THEN 1 ELSE 0 END) AS osm_matched,
              SUM(CASE WHEN t.osi_id IS NOT NULL AND t.osi_id<>'' THEN 1 ELSE 0 END) AS osi_matched,
              SUM(CASE WHEN t.vrti_id IS NOT NULL AND t.vrti_id<>'' THEN 1 ELSE 0 END) AS vrti_matched,
              SUM(CASE WHEN t.logainm_id IS NOT NULL AND t.logainm_id<>'' THEN 1 ELSE 0 END) AS logainm_matched
            FROM townland t
            WHERE t.name IN (SELECT DISTINCT townland_norm FROM unified_record WHERE townland_norm IS NOT NULL)
            """
        ).fetchone()
    )


def main() -> int:
    from create_app import create_app
    from extensions import get_db_conn

    app = create_app()
    with app.app_context():
        conn = get_db_conn()

        print("=" * 100)
        print("1. RAW SOURCE FILE ROW COUNTS")
        print("=" * 100)
        raw = raw_source_counts()
        for k, v in raw.items():
            print(f"  {k}: {v}")

        print()
        print("=" * 100)
        print("2. CURRENTLY-LOADED DB ROW COUNTS")
        print("=" * 100)
        loaded = loaded_db_counts(conn)
        for k, v in loaded.items():
            print(f"  {k}: {v}")

        print()
        print("=" * 100)
        print("3. INGESTION COMPLETENESS (raw vs loaded, where directly comparable)")
        print("=" * 100)
        print(f"  unified_processed.csv ({raw['unified_processed.csv']}) -> unified_record ({loaded['unified_record']}): "
              f"{'MATCH — 100% complete' if raw['unified_processed.csv'] == loaded['unified_record'] else 'MISMATCH'}")
        print(f"  workhouse_data_final.xlsx total ({raw['workhouse_data_final.xlsx[TOTAL]']}) -> source_mentions ({loaded['source_mentions']}): "
              f"{'MATCH — 100% complete' if raw['workhouse_data_final.xlsx[TOTAL]'] == loaded['source_mentions'] else 'MISMATCH'}")
        print(f"  townlands.json features ({raw['townlands.json[features]']}) -> estate townland universe: "
              "see full_ingest dry-run below, not a direct 1:1 table count")
        print("  census_record and clearances_record are NOT a direct 1:1 file-to-table mapping —")
        print("  census_record blends live VRTI KG (1841-1891) + GeoJSON-embedded + CSV-seed fallback")
        print("  (1827-1868); clearances_record comes from GeoJSON-embedded per-townland fields.")
        print("  See the full_ingest dry-run below for what each source currently contributes.")

        print()
        print("=" * 100)
        print("4. FULL_INGEST DRY RUN (safe — writes nothing) vs CURRENT DB STATE")
        print("=" * 100)
        from backend.jobs.full_ingest import run_full_ingest
        stats = run_full_ingest(dry_run=True)
        print(f"  Dry-run stats: {stats}")
        print()
        print(f"  townlands.json features: 152, processed: {stats['townlands_processed']}, "
              f"KG-enriched (any geometry/centroid): {stats['townlands_kg_enriched']}")
        print(f"  census_records_kg (live VRTI fetch, right now): {stats['census_records_kg']}")
        print(f"  census_records_json (GeoJSON-embedded): {stats['census_records_json']}")
        print(f"  census_records_csv_seed (estate fallback, used because KG returned 0): {stats['census_records_csv_seed']}")
        fresh_census_total = stats['census_records_kg'] + stats['census_records_json'] + stats['census_records_csv_seed']
        print(f"  Fresh-rebuild census total: {fresh_census_total}  vs  currently stored: {loaded['census_record']}")
        print(f"  clearances_records (fresh dry run): {stats['clearances_records']}  vs  currently stored: {loaded['clearances_record']}")
        print()
        if stats['census_records_kg'] == 0:
            print("  *** WARNING: live VRTI census KG endpoint returned 0 rows just now. ***")
            print("  *** A real (non-dry-run) full_ingest today would SHRINK census_record from")
            print(f"  *** {loaded['census_record']} to ~{fresh_census_total} rows. DO NOT run a real rebuild")
            print("  *** against this database until this is investigated — see evidence doc.")

        print()
        print("=" * 100)
        print("5. TOWNLAND ALIGNMENT COVERAGE (scoped to ~152 estate townlands)")
        print("=" * 100)
        align = alignment_coverage(conn)
        print(f"  {align}")

        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
