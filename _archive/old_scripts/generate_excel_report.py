import pandas as pd
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.append(os.getcwd())

from coolattin.services.fuzzy import norm_key, _score

def generate_excel_report():
    input_csv = "coolattin/static/data/matched_records.csv"
    output_xlsx = "coolattin/static/data/mapping_audit.xlsx"
    
    print(f"Loading data from {input_csv}...")
    try:
        df = pd.read_csv(input_csv)
    except Exception as e:
        print(f"Error loading CSV: {e}")
        return

    print("Generating Excel report...")
    wb = Workbook()
    
    ws_audit = wb.active
    ws_audit.title = "Mapping Audit"
    
    headers = [
        "Cluster (Canon)", "Raw Surname", "Forename", "Year", "Townland", 
        "Source", "ID", "Norm Key", "Match Score", "Notes"
    ]
    ws_audit.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws_audit[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    df = df.sort_values(["surname_canon", "townland_norm", "year_norm"])

    for _, row in df.iterrows():
        canon = str(row.get("surname_canon", ""))
        raw = str(row.get("surname_raw", ""))
        fr = str(row.get("forename_raw", ""))
        yr = str(row.get("year_norm", ""))
        town = str(row.get("townland_norm", ""))
        src = str(row.get("source", ""))
        rid = str(row.get("id", ""))
        
        norm = norm_key(raw)
        norm_canon = norm_key(canon)
        
        score = _score(norm_canon, norm)
        
        notes = ""
        if raw != canon:
            notes = f"Variation of {canon}"
        if score < 100:
            notes += f" (Fuzzy match: {score})"

        ws_audit.append([
            canon, raw, fr, yr, town, src, rid, norm, score, notes
        ])

    ws_summary = wb.create_sheet("Cluster Summary")
    ws_summary.append(["Cluster (Canon)", "Record Count", "Unique Raw Surnames", "Sources involved"])
    
    summary_data = df.groupby("surname_canon").agg({
        "id": "count",
        "surname_raw": lambda x: ", ".join(x.unique()),
        "source": lambda x: ", ".join(x.unique())
    }).reset_index()
    
    for _, row in summary_data.iterrows():
        ws_summary.append(row.tolist())

    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill

    for sheet in [ws_audit, ws_summary]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = min(adjusted_width, 50)

    try:
        wb.save(output_xlsx)
        print(f"Excel report generated at: {output_xlsx}")
    except Exception as e:
        print(f"Error saving Excel: {e}")

if __name__ == "__main__":
    generate_excel_report()
