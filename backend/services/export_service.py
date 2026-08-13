from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)


def export_census(
    records: list,
    scope,
    extra_meta: Optional[dict] = None,
) -> str:
    try:
        import openpyxl
    except ImportError:
        log.error("export_service.missing_openpyxl — install openpyxl to enable exports")
        raise

    from backend.config import ActiveConfig
    from backend.integrations.vrti_sparql import SPARQL_ENDPOINT

    exports_dir = ActiveConfig.EXPORTS_DIR / "census"
    exports_dir.mkdir(parents=True, exist_ok=True)

    year_part = f"_{scope.year}" if hasattr(scope, "year") and scope.year else "_all"
    if isinstance(scope, dict) and scope.get("year"):
        year_part = f"_{scope['year']}"
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"census_wicklow{year_part}_{timestamp}.xlsx"
    filepath = exports_dir / filename

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Census Records"
    headers = [
        "Townland", "Year", "Male", "Female", "Total",
        "Inhabited Houses", "Uninhabited Houses", "Source", "KG URI",
    ]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        if hasattr(rec, "to_dict"):
            r = rec.to_dict()
        else:
            r = rec
        ws.append([
            r.get("townland") or r.get("townland_name", ""),
            r.get("year"),
            r.get("male"),
            r.get("female"),
            r.get("total"),
            r.get("inhabited"),
            r.get("uninhabited"),
            r.get("source", ""),
            r.get("kg_uri", "") or "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    meta_ws = wb.create_sheet("Export Metadata")
    meta_ws.column_dimensions["A"].width = 28
    meta_ws.column_dimensions["B"].width = 60

    scope_str = str(vars(scope)) if hasattr(scope, "__dict__") else str(scope)
    rows_meta = [
        ("Field", "Value"),
        ("Generated At (UTC)", datetime.now(timezone.utc).isoformat()),
        ("Source Endpoint", SPARQL_ENDPOINT),
        ("Query Scope", scope_str),
        ("Record Count", len(records)),
        ("Export File", str(filepath)),
        ("Application", "Coolattin Lineage — Digital Estate Archive"),
        ("Census Years Covered", "1841, 1851, 1861, 1871, 1881, 1891"),
        ("County", "Wicklow, Ireland"),
    ]
    if extra_meta:
        for k, v in extra_meta.items():
            rows_meta.append((k, str(v)))

    for i, (k, v) in enumerate(rows_meta):
        meta_ws.append([k, v])
        if i == 0:
            for cell in meta_ws[1]:
                cell.font = Font(bold=True)

    wb.save(filepath)
    log.info(
        "export_service.census_written | path=%s records=%d",
        filepath, len(records),
    )
    return str(filepath)


def export_townlands(townlands: list, extra_meta: Optional[dict] = None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.error("export_service.missing_openpyxl")
        raise

    from backend.config import ActiveConfig

    exports_dir = ActiveConfig.EXPORTS_DIR / "townlands"
    exports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"townlands_wicklow_{timestamp}.xlsx"
    filepath = exports_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wicklow Townlands"

    headers = [
        "Name", "Gaelic Name", "Barony", "Civil Parish",
        "Electoral Division", "Placename Theme", "Description",
        "KG URI", "Source",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")

    for t in townlands:
        if hasattr(t, "to_dict"):
            r = t.to_dict()
        else:
            r = t
        ws.append([
            r.get("name", ""),
            r.get("name_gaelic", ""),
            r.get("barony", ""),
            r.get("civil_parish", ""),
            r.get("electoral_division", ""),
            r.get("placename_theme", ""),
            r.get("description", ""),
            r.get("kg_uri", ""),
            r.get("source", ""),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    meta_ws = wb.create_sheet("Export Metadata")
    meta_ws.append(["Field", "Value"])
    meta_ws.append(["Generated At (UTC)", datetime.now(timezone.utc).isoformat()])
    meta_ws.append(["Record Count", len(townlands)])
    meta_ws.append(["County", "Wicklow, Ireland"])
    meta_ws.append(["Application", "Coolattin Lineage — Digital Estate Archive"])
    if extra_meta:
        for k, v in extra_meta.items():
            meta_ws.append([k, str(v)])

    wb.save(filepath)
    log.info("export_service.townlands_written | path=%s count=%d", filepath, len(townlands))
    return str(filepath)


def get_latest_census_export() -> dict:
    from backend.repositories import refresh_state_repository
    state = refresh_state_repository.get("wicklow_census")
    if not state or not state.export_file:
        return {"export_file": None, "message": "No export available yet. Run a refresh first."}
    path = Path(state.export_file)
    return {
        "export_file": state.export_file,
        "exists": path.exists(),
        "generated_at": state.last_synced_at,
        "record_count": state.record_count,
        "source": state.source,
    }


def regenerate_from_db(year: Optional[int] = None) -> str:
    from backend.repositories import census_repository
    from backend.models.census_models import CensusFilters

    filters = CensusFilters(year=year, limit=10000)
    records = census_repository.find(filters)
    return export_census([r.to_dict() for r in records], filters,
                         extra_meta={"Export Type": "regenerated_from_db"})
